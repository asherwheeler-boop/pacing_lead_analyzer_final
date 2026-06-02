from __future__ import annotations
from io import BytesIO
from pathlib import Path
import numpy as np
import pandas as pd


def load_builtin_dd_dataset(dataset_label: str):
    mapping = {
        "Patient 1011": "Patient_1011_DD0102.xlsx",
        "Patient 1012": "Patient_1012_DD0102.xlsx",
    }
    file_name = mapping.get(dataset_label)
    if not file_name:
        raise ValueError(f"Unsupported dataset label: {dataset_label}")
    path = Path(file_name)
    if not path.exists():
        raise FileNotFoundError(f"Reference file not found: {file_name}")
    df = pd.read_excel(path, engine="openpyxl")
    return _normalize_dataset(df), file_name


def load_user_dataset(file_obj):
    name = getattr(file_obj, "name", "")
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    if name.lower().endswith((".xlsx", ".xls")):
        engine = "openpyxl" if name.lower().endswith(".xlsx") else "xlrd"
        df = pd.read_excel(file_obj, engine=engine)
    else:
        df = pd.read_csv(file_obj)
    return _normalize_dataset(df)


def _normalize_dataset(df: pd.DataFrame) -> pd.DataFrame:
    cols = {str(c).strip().lower(): c for c in df.columns}

    def pick(*aliases):
        for alias in aliases:
            if alias.lower() in cols:
                return cols[alias.lower()]
        return None

    segment_col = pick("Segment", "Lead section", "Lead Section", "segment")
    re_col = pick("Re (mm)", "Re", "Re [mm]", "re (mm)", "re")
    ri_col = pick("Ri (mm)", "Ri", "Ri [mm]", "ri (mm)", "ri")
    ca_col = pick("Ca (1/cm)", "Ca (/cm)", "Ca", "Ca [1/cm]", "ca (1/cm)", "ca")

    if segment_col is None or ca_col is None:
        raise ValueError("Dataset must contain at least Segment and Ca columns.")

    out = pd.DataFrame({
        "Segment": pd.to_numeric(df[segment_col], errors="coerce"),
        "Re (mm)": pd.to_numeric(df[re_col], errors="coerce") if re_col else np.nan,
        "Ri (mm)": pd.to_numeric(df[ri_col], errors="coerce") if ri_col else np.nan,
        "Ca (1/cm)": pd.to_numeric(df[ca_col], errors="coerce"),
    })
    out = out.dropna(subset=["Segment", "Ca (1/cm)"]).copy()
    out["Segment"] = out["Segment"].astype(int)
    out = out.sort_values("Segment").reset_index(drop=True)
    return out


def compare_datasets(user_df: pd.DataFrame, dd_df: pd.DataFrame, dataset_label: str):
    merged = pd.merge(
        dd_df,
        user_df,
        on="Segment",
        how="inner",
        suffixes=("_dd", "_user")
    )
    if merged.empty:
        return pd.DataFrame(), {"n_segments": 0, "mae": np.nan, "rmse": np.nan, "max_abs_error": np.nan}

    merged["Difference"] = merged["Ca (1/cm)_user"] - merged["Ca (1/cm)_dd"]
    merged["Abs_Error"] = merged["Difference"].abs()

    def classify(x):
        if pd.isna(x):
            return "N/A"
        if x <= 0.02:
            return "PASS"
        if x <= 0.05:
            return "WARNING"
        return "FAIL"

    merged["Status"] = merged["Abs_Error"].apply(classify)

    compare_df = pd.DataFrame({
        "Segment": merged["Segment"],
        "DD0102_Re (mm)": merged["Re (mm)_dd"],
        "DD0102_Ri (mm)": merged["Ri (mm)_dd"],
        "DD0102_Ca": merged["Ca (1/cm)_dd"],
        "User_Re (mm)": merged["Re (mm)_user"],
        "User_Ri (mm)": merged["Ri (mm)_user"],
        "User_Ca": merged["Ca (1/cm)_user"],
        "Difference": merged["Difference"],
        "Abs_Error": merged["Abs_Error"],
        "Status": merged["Status"],
    }).sort_values("Segment").reset_index(drop=True)

    metrics = {
        "n_segments": int(len(compare_df)),
        "mae": float(compare_df["Abs_Error"].mean()),
        "rmse": float(np.sqrt(np.mean(compare_df["Difference"] ** 2))),
        "max_abs_error": float(compare_df["Abs_Error"].max()),
    }
    return compare_df, metrics


def generate_comparison_plot(compare_df: pd.DataFrame, dataset_label: str):
    try:
        import plotly.graph_objects as go
    except ImportError:
        return None
    if compare_df is None or compare_df.empty:
        return None
    color_map = {"PASS": "#2ecc71", "WARNING": "#f1c40f", "FAIL": "#e74c3c", "N/A": "#95a5a6"}
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=compare_df["Segment"],
        y=compare_df["User_Ca"],
        mode="lines+markers",
        name="User Ca",
        line=dict(width=3),
        marker=dict(color=[color_map.get(s, "#95a5a6") for s in compare_df["Status"]], size=9),
    ))
    fig.add_trace(go.Scatter(
        x=compare_df["Segment"],
        y=compare_df["DD0102_Ca"],
        mode="lines+markers",
        name=f"{dataset_label} DD-0102 Ca",
        line=dict(width=3, dash="dash"),
    ))
    fig.update_layout(
        title=f"User dataset vs {dataset_label}",
        xaxis_title="Segment",
        yaxis_title="Ca (1/cm)",
        hovermode="x unified",
        plot_bgcolor="white",
        paper_bgcolor="white",
        height=520,
    )
    return fig


def export_comparison_excel(compare_df: pd.DataFrame, metrics: dict, dataset_label: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    ws.append(["Reference dataset", dataset_label])
    ws.append(["Segments compared", metrics.get("n_segments")])
    ws.append(["Mean absolute error", metrics.get("mae")])
    ws.append(["RMSE", metrics.get("rmse")])
    ws.append(["Largest absolute error", metrics.get("max_abs_error")])
    ws.append([])
    headers = list(compare_df.columns)
    ws.append(headers)
    for row in compare_df.itertuples(index=False):
        ws.append(list(row))

    # format header
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[7]:
        cell.fill = header_fill
        cell.font = header_font

    # color status column
    status_col = headers.index("Status") + 1
    fills = {
        "PASS": PatternFill("solid", fgColor="D4EDDA"),
        "WARNING": PatternFill("solid", fgColor="FFF3CD"),
        "FAIL": PatternFill("solid", fgColor="F8D7DA"),
    }
    for r in range(8, ws.max_row + 1):
        val = ws.cell(r, status_col).value
        if val in fills:
            ws.cell(r, status_col).fill = fills[val]

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()
